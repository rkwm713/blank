from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# Import shared utility to avoid duplicated stub implementation
from trace_utils import get_trace_by_id
from utils import inches_to_feet_inches_str, extract_string_value
from wire_utils import parse_feet_inches_str_to_inches as feet_inches_str_to_inches

def categorize_wire(wire_type):
    """Categorize wire as COMM, CPS, or OTHER based on description."""
    if not wire_type:
        return 'OTHER'
        
    wire_type_lower = str(wire_type).lower()
    
    if any(comm in wire_type_lower for comm in ['fiber', 'comm', 'telco', 'spectrum', 'at&t']):
        return 'COMM'
    elif any(cps in wire_type_lower for cps in ['cps', 'neutral', 'primary', 'secondary']):
        return 'CPS'
    else:
        return 'OTHER'

def find_lowest_height(wires):
    """Find the wire with the lowest height from a list."""
    if not wires:
        return None
        
    lowest = wires[0][1]  # Start with first wire
    
    for _, height_data in wires[1:]:
        if height_data['height'] < lowest['height']:
            lowest = height_data
            
    return lowest

def find_proposed_for_category(proposed_heights, category):
    """Find proposed height for a wire category."""
    lowest_proposed = None
    
    for wire_type, height_data in proposed_heights.items():
        if categorize_wire(wire_type) == category:
            if lowest_proposed is None or height_data['height'] < lowest_proposed['height']:
                lowest_proposed = height_data
                
    return lowest_proposed

def process_connection_heights(pole_data):
    """Extract lowest wire heights from all sections of a connection."""
    processed_heights = {}
    
    # Get all connections for this pole
    for connection in pole_data.get('connections', []):
        connection_id = connection.get('connection_id')
        if not connection_id:
            continue
            
        # Process all sections in this connection
        wire_heights = {}  # To track lowest height per wire type
        
        for section_id, section in connection.get('sections', {}).items():
            # Extract all wires from this section's photos
            for photo_id, photo_data in section.get('photos', {}).items():
                # Get wire data from the photo
                wires = []
                photofirst_data = photo_data.get('photofirst_data', {})
                
                # Handle wire data as either list or dictionary
                wire_data = photofirst_data.get('wire', [])
                if isinstance(wire_data, list):
                    wires = wire_data
                elif isinstance(wire_data, dict):
                    wires = list(wire_data.values())
                
                # Process each wire
                for wire in wires:
                    if not isinstance(wire, dict):
                        continue
                    
                    # Get trace data for wire description
                    trace_id = wire.get('_trace', '')
                    trace_data = get_trace_by_id(pole_data, trace_id)
                    
                    # Create a description from trace data
                    wire_description = ''
                    if trace_data and isinstance(trace_data, dict):
                        owner = trace_data.get('company', '')
                        cable_type = trace_data.get('cable_type', '')
                        wire_description = f"{owner} {cable_type}".strip()
                    
                    # Get measured height
                    height = wire.get('_measured_height')
                    
                    if not height or not wire_description:
                        continue
                    
                    # Convert to float for comparison
                    try:
                        height_value = float(height)
                        
                        # Store lowest height for this wire type
                        if wire_description not in wire_heights or height_value < wire_heights[wire_description]['height']:
                            wire_heights[wire_description] = {
                                'height': height_value,
                                'height_str': inches_to_feet_inches_str(height_value),
                                'wire_data': wire,
                                'trace_data': trace_data
                            }
                    except (ValueError, TypeError):
                        pass
        
        # Store the results
        processed_heights[connection_id] = wire_heights
        
    return processed_heights

def determine_proposed_heights(pole_data, connection_heights):
    """
    Calculate proposed heights for each connection and wire type.
    
    Args:
        pole_data (dict): Pole data dictionary
        connection_heights (dict): Dictionary of connection heights
        
    Returns:
        dict: Dictionary mapping connection_id -> wire_type -> {height, height_str}
    """
    proposed_heights = {}
    
    for connection_id, wire_heights in connection_heights.items():
        proposed_heights[connection_id] = {}
        
        for wire_type, height_data in wire_heights.items():
            wire = height_data.get('wire_data', {})
            
            # Check if this wire has an effective move or MR move value
            # This is a simplified check - actual implementation would need to 
            # look at the specific fields in your data model
            has_move = (
                wire.get('effective_move') or 
                wire.get('mr_move') or
                height_data.get('trace_data', {}).get('proposed', False)
            )
            
            if has_move:
                # In a real implementation, you would calculate the proposed height
                # based on your specific data model and business rules
                # For now, we'll just use a placeholder
                proposed_height = wire.get('_proposed_height')
                
                if proposed_height:
                    try:
                        proposed_height_value = float(proposed_height)
                        proposed_heights[connection_id][wire_type] = {
                            'height': proposed_height_value,
                            'height_str': inches_to_feet_inches_str(proposed_height_value)
                        }
                    except (ValueError, TypeError):
                        pass
    
    return proposed_heights

def identify_ref_subgroups(pole_data):
    """Identify reference subgroups that need special handling."""
    ref_groups = []
    
    # First check the reference_spans data that should be properly processed already
    for ref_span in pole_data.get('reference_spans', []):
        if isinstance(ref_span, dict) and 'header' in ref_span:
            header = ref_span.get('header', {})
            ref_groups.append({
                'connection_id': ref_span.get('connection_id', ''),
                'label': header.get('description', ''),
                'style_hint': header.get('style_hint', 'orange'),
                'from_pole': pole_data.get('pole_number', ''),
                'to_pole': header.get('to_pole', ''),
                'attachments': ref_span.get('attachments', [])
            })
    
    # If we don't have any reference_spans data, fall back to connections
    if not ref_groups:
        for connection in pole_data.get('connections', []):
            # Check if this is a reference subgroup
            is_ref = False
            description = connection.get('to_pole', '')
            
            # Check for common patterns in service pole descriptions
            if description and any(keyword in str(description).lower() 
                                for keyword in ['service', 'red', 'north', 'south', 'east', 'west']):
                is_ref = True
            
            if is_ref:
                # Try to extract direction from description or use a default
                direction = "North East"  # Default direction
                
                # Simple direction extraction - can be enhanced based on actual data patterns
                if "north" in str(description).lower() and "east" in str(description).lower():
                    direction = "North East"
                elif "north" in str(description).lower() and "west" in str(description).lower():
                    direction = "North West"
                elif "south" in str(description).lower() and "east" in str(description).lower():
                    direction = "South East"
                elif "south" in str(description).lower() and "west" in str(description).lower():
                    direction = "South West"
                
                # Create the label in the format shown in the expected image
                ref_groups.append({
                    'connection_id': connection.get('connection_id'),
                    'label': f"Ref ({direction}) to {description}",
                    'style_hint': 'orange',  # Default to orange for normal references
                    'from_pole': connection.get('from_pole'),
                    'to_pole': description,
                    'attachments': []  # Will be populated later
                })
    
    return ref_groups

def create_make_ready_excel(processed_pole_data, output_filepath):
    """
    Generate Excel report from processed pole data.
    
    Args:
        processed_pole_data (list): List of processed pole dictionaries
        output_filepath (str): Output Excel file path
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Make Ready Report"
    
    # ===== Setup Styles =====
    # Define styles for different cell types
    header_font = Font(bold=True)
    
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    purple_fill = PatternFill(start_color="E6C3E6", end_color="E6C3E6", fill_type="solid")
    header_fill = PatternFill(start_color="A6C9EC", end_color="A6C9EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== Set Column Widths =====
    column_widths = {
        'A': 10,  # Operation Number
        'B': 15,  # Attachment Action
        'C': 15,  # Pole Owner
        'D': 12,  # Pole #
        'E': 20,  # Pole Structure
        'F': 15,  # Proposed Riser
        'G': 15,  # Proposed Guy
        'H': 15,  # PLA
        'I': 15,  # Construction Grade
        'J': 15,  # Height Lowest Com
        'K': 15,  # Height Lowest CPS Elec
        'L': 25,  # Attacher Description
        'M': 12,  # Existing (Attachment Height)
        'N': 12,  # Proposed (Attachment Height)
        'O': 12,  # Mid-Span
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ===== Create Headers =====
    # Row 1: Merged headers across columns
    ws.cell(row=1, column=10).value = "Existing Mid-Span Data"
    ws.merge_cells('J1:K1')
    ws.cell(row=1, column=10).fill = header_fill
    ws.cell(row=1, column=10).font = header_font
    ws.cell(row=1, column=10).alignment = center_alignment
    
    ws.cell(row=1, column=12).value = "Make Ready Data"
    ws.merge_cells('L1:O1')
    ws.cell(row=1, column=12).fill = header_fill
    ws.cell(row=1, column=12).font = header_font
    ws.cell(row=1, column=12).alignment = center_alignment
    
    # Define header values for columns A-I
    main_headers = [
        "Operation Number", 
        "Attachment Action\n(I)nstalling\n(R)emoving\n(E)xisting", 
        "Pole Owner", 
        "Pole #", 
        "Pole Structure", 
        "Proposed Riser\n(Yes/No) &\nQTY", 
        "Proposed Guy\n(Yes/No) &\nQTY", 
        "PLA (%) with proposed attachment", 
        "Construction Grade of Analysis"
    ]
    
    # Define header values for columns J-O in row 2
    detail_headers = [
        "Height Lowest Com", 
        "Height Lowest CPS Electrical",
        "Attacher Description",
        "Attachment Height",  # This will be merged M2:N2
        "", # Skip cell N2 as it will be merged
        "Mid-Span\n(same span as existing)\nProposed"
    ]
    
    # Set values and merge columns A-I vertically (rows 1-3)
    for col_idx, header in enumerate(main_headers, start=1):
        # Set value in the top cell (row 1)
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Merge vertically (rows 1-3)
        ws.merge_cells(start_row=1, start_column=col_idx, 
                      end_row=3, end_column=col_idx)
    
    # Set values for columns J-O in row 2
    for col_idx, header in enumerate(detail_headers, start=10):
        if col_idx != 14:  # Skip the column that will be part of the merge
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
    
    # Merge columns J, K, and L vertically across rows 2-3
    ws.merge_cells('J2:J3')  # Merge "Height Lowest Com" across rows 2-3
    ws.merge_cells('K2:K3')  # Merge "Height Lowest CPS Electrical" across rows 2-3
    ws.merge_cells('L2:L3')  # Merge "Attacher Description" across rows 2-3
    
    # Set up merged header for Attachment Height (after setting values)
    ws.merge_cells('M2:N2')
    
    # Add subheaders for Attachment Height
    ws.cell(row=3, column=13).value = "Existing"
    ws.cell(row=3, column=13).fill = header_fill
    ws.cell(row=3, column=13).font = header_font
    ws.cell(row=3, column=13).alignment = center_alignment
    ws.cell(row=3, column=13).border = thin_border
    
    ws.cell(row=3, column=14).value = "Proposed"
    ws.cell(row=3, column=14).fill = header_fill
    ws.cell(row=3, column=14).font = header_font
    ws.cell(row=3, column=14).alignment = center_alignment
    ws.cell(row=3, column=14).border = thin_border
    
    # Set "Proposed" in column O row 3
    ws.cell(row=3, column=15).value = "Proposed"
    ws.cell(row=3, column=15).fill = header_fill
    ws.cell(row=3, column=15).font = header_font
    ws.cell(row=3, column=15).alignment = center_alignment
    ws.cell(row=3, column=15).border = thin_border
    
    # Freeze panes to keep headers visible
    ws.freeze_panes = 'A4'  # Now shifted down one row to account for subheaders
    
    # ===== Populate Data =====
    current_row = 4  # Start one row lower to account for subheaders
    operation_number = 1
    
    # First pass: process only primary operation poles (those in SPIDAcalc)
    for pole in processed_pole_data:
        # Skip non-primary poles on the first pass
        if not pole.get('is_primary', False):
            continue
            
        print(f"[DEBUG] Processing primary pole: {pole.get('pole_number')}")
        
        # Consistently use pole.get('attachers', []) as the primary source for rendering items in columns L-O.
        # It's assumed that any necessary filtering or ordering (like "attachments below neutral")
        # has been handled upstream and incorporated into this 'attachers' list.
        attachers = pole.get('attachers', [])
        
        # Calculate how many rows this pole will span
        # Start with attachers count and add rows for midspan data and separators
        attachers_count = len(attachers)
        
        # Process connection data for this pole
        connection_heights = process_connection_heights(pole)
        proposed_heights = determine_proposed_heights(pole, connection_heights)
        ref_groups = identify_ref_subgroups(pole)
        
        # Calculate total rows needed:
        # - Attachers rows
        # - Midspan data (lowest com/cps values in J/K)
        # - From/To pole header and data (2 rows)
        # - REF groups (header + items)
        ref_group_rows = 0
        for ref in ref_groups:
            # Each REF group has a header row
            ref_group_rows += 1
            # Plus rows for each connection's wire types
            connection_id = ref.get('connection_id')
            if connection_id in connection_heights:
                ref_group_rows += len(connection_heights[connection_id])
        
        # Add 1 for the midspan values that appear before the From/To pole headers
        total_rows = max(1, attachers_count) + 3 + ref_group_rows
        
        # Start row for this pole
        pole_start_row = current_row
        pole_end_row = pole_start_row + total_rows - 1
        
        # ===== Write Pole-Level Data (Columns A-I) =====
        # Apply header style to all cells in the operation header row
        for col in range(1, 10):
            cell = ws.cell(row=pole_start_row, column=col)
            cell.fill = header_fill  # Apply the blue header fill
            cell.font = header_font  # Apply bold text for headers
            
            # Merge cells for pole-level data if spanning multiple rows
            if total_rows > 1:
                ws.merge_cells(start_row=pole_start_row, start_column=col, 
                              end_row=pole_end_row, end_column=col)
            
            # Apply styling for merged cells
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # Fill in pole data - using extract_string_value to safely handle potential dictionaries
        ws.cell(row=pole_start_row, column=1).value = operation_number  # Operation Number (sequential for primary poles)
        ws.cell(row=pole_start_row, column=2).value = extract_string_value(pole.get('pole_action', '(I)nstalling'))
        ws.cell(row=pole_start_row, column=3).value = "CPS"  # Per spec, always CPS
        ws.cell(row=pole_start_row, column=4).value = extract_string_value(pole.get('pole_number'))
        ws.cell(row=pole_start_row, column=5).value = extract_string_value(pole.get('pole_structure'))
        
        # Make sure the proposed_riser and proposed_guy values are correctly formatted and capitalized
        proposed_riser = extract_string_value(pole.get('proposed_riser', 'NO'))
        proposed_guy = extract_string_value(pole.get('proposed_guy', 'NO'))
        pla_percentage = extract_string_value(pole.get('pla_percentage', 'N/A'))
        
        # Debug output to see what's being processed
        print(f"[DEBUG] Excel Values - Primary Pole: {pole.get('pole_number')}, Proposed Guy: {proposed_guy}, PLA: {pla_percentage}")
        
        # Ensure consistent capitalization
        if proposed_riser.lower() == 'no':
            proposed_riser = 'NO'
        if proposed_guy.lower() == 'no':
            proposed_guy = 'NO'
            
        # Write cell values
        ws.cell(row=pole_start_row, column=6).value = proposed_riser
        ws.cell(row=pole_start_row, column=7).value = proposed_guy
        ws.cell(row=pole_start_row, column=8).value = pla_percentage
        ws.cell(row=pole_start_row, column=9).value = extract_string_value(pole.get('construction_grade'))
        
        # ===== Write Midspan Data (J-K) =====
        # For existing midspan data cell (Columns J/K): use new midspan_heights helper
        from_pole = extract_string_value(pole.get('from_pole'))
        to_pole = extract_string_value(pole.get('to_pole'))
        midspan_heights = pole.get('midspan_heights', {})
        comm_val = 'NA'
        cps_val = 'NA'
        # Try both directions (from->to or to->from)
        if midspan_heights:
            # Try (from_pole, to_pole) as a tuple key
            key_tuple = (from_pole, to_pole)
            key_tuple_rev = (to_pole, from_pole)
            # Try as string keys (legacy)
            if key_tuple in midspan_heights:
                comm_val = midspan_heights[key_tuple].get('comm', 'NA')
                cps_val = midspan_heights[key_tuple].get('cps', 'NA')
            elif key_tuple_rev in midspan_heights:
                comm_val = midspan_heights[key_tuple_rev].get('comm', 'NA')
                cps_val = midspan_heights[key_tuple_rev].get('cps', 'NA')
            elif to_pole in midspan_heights:
                comm_val = midspan_heights[to_pole].get('comm', 'NA')
                cps_val = midspan_heights[to_pole].get('cps', 'NA')
            elif from_pole in midspan_heights:
                comm_val = midspan_heights[from_pole].get('comm', 'NA')
                cps_val = midspan_heights[from_pole].get('cps', 'NA')
        lowest_com_cell = ws.cell(row=pole_start_row, column=10)
        lowest_com_cell.value = comm_val
        lowest_com_cell.alignment = center_alignment
        lowest_com_cell.border = thin_border

        lowest_cps_cell = ws.cell(row=pole_start_row, column=11)
        lowest_cps_cell.value = cps_val
        lowest_cps_cell.alignment = center_alignment
        lowest_cps_cell.border = thin_border
        
        # ===== Write Attachers Data (Columns L-O) =====
        attacher_row = pole_start_row
        for attacher in attachers:
            # Check if this is a reference or backspan header
            attacher_type = attacher.get('type', '')
            if attacher_type in ['reference_header', 'backspan_header']:
                # Special handling for reference and backspan headers
                header_text = attacher.get('description', '')
                style_hint = attacher.get('style_hint', '')
                
                # Create header cell with merged cells L-O
                header_cell = ws.cell(row=attacher_row, column=12, value=header_text)
                
                # Determine fill color based on style_hint
                if style_hint == 'orange':
                    header_fill_color = orange_fill
                elif style_hint == 'purple':
                    header_fill_color = purple_fill
                elif style_hint == 'light-blue':
                    header_fill_color = light_blue_fill
                else:
                    # If header text contains South East, use purple
                    if 'south east' in header_text.lower() or 'southeast' in header_text.lower():
                        header_fill_color = purple_fill
                    else:
                        header_fill_color = orange_fill  # Default to Orange
                
                # Apply styles
                header_cell.fill = header_fill_color
                header_cell.font = header_font
                header_cell.alignment = center_alignment
                header_cell.border = thin_border
                
                # Merge cells L-O
                ws.merge_cells(start_row=attacher_row, start_column=12, 
                              end_row=attacher_row, end_column=15)
                
                # Add borders to all merged cells
                for col in range(12, 16):
                    cell = ws.cell(row=attacher_row, column=col)
                    cell.border = thin_border
                
                attacher_row += 1
                continue  # Skip to next attacher
            
            # Regular attacher data - safely extract string values
            desc_cell = ws.cell(row=attacher_row, column=12)
            desc_cell.value = extract_string_value(attacher.get('description'))
            
            # Extract attachment heights
            existing_height = extract_string_value(attacher.get('existing_height'))
            proposed_height = extract_string_value(attacher.get('proposed_height'))
            midspan_proposed = extract_string_value(attacher.get('midspan_proposed'))

            # Set the cell values
            ws.cell(row=attacher_row, column=13).value = existing_height
            ws.cell(row=attacher_row, column=14).value = proposed_height if proposed_height not in (None, '', 'N/A') else ''
            ws.cell(row=attacher_row, column=15).value = midspan_proposed
            
            # Apply border and alignment
            for col in range(12, 16):
                cell = ws.cell(row=attacher_row, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Left-align the attacher description
            desc_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            attacher_row += 1
        
        # If no attachers, add at least one empty row
        if not attachers:
            for col in range(12, 16):
                cell = ws.cell(row=attacher_row, column=col)
                cell.border = thin_border
            attacher_row += 1
        
        # ===== Write From/To Pole Section =====
        # Now add the From/To pole headers AFTER the midspan data
        midspan_row = attacher_row
        
        # Now merge the midspan data cells vertically from pole_start_row to (midspan_row-1)
        if midspan_row > pole_start_row:
            ws.merge_cells(start_row=pole_start_row, start_column=10, 
                          end_row=midspan_row-1, end_column=10)
            ws.merge_cells(start_row=pole_start_row, start_column=11, 
                          end_row=midspan_row-1, end_column=11)
        
        # From/To Pole Header Row
        from_cell = ws.cell(row=midspan_row, column=10)
        from_cell.value = "From Pole"
        from_cell.fill = light_blue_fill
        from_cell.font = header_font
        from_cell.alignment = center_alignment
        from_cell.border = thin_border
        
        to_cell = ws.cell(row=midspan_row, column=11)
        to_cell.value = "To Pole"
        to_cell.fill = light_blue_fill
        to_cell.font = header_font
        to_cell.alignment = center_alignment
        to_cell.border = thin_border
        
        # From/To Pole Data Row
        midspan_row += 1
        from_pole_cell = ws.cell(row=midspan_row, column=10)
        from_pole_cell.value = extract_string_value(pole.get('from_pole'))
        from_pole_cell.alignment = center_alignment
        from_pole_cell.border = thin_border
        
        to_pole_cell = ws.cell(row=midspan_row, column=11)
        to_pole_cell.value = extract_string_value(pole.get('to_pole'))
        to_pole_cell.alignment = center_alignment
        to_pole_cell.border = thin_border
        
        # Add proposed midspan height in column O if available
        proposed_midspan_cell = ws.cell(row=midspan_row, column=15)
        
        # Add empty cells with borders for columns L-N
        for col in range(12, 15):
            cell = ws.cell(row=midspan_row, column=col)
            cell.border = thin_border
        
        # The From/To pole row should only show midspan_proposed for the overall span
        # (not attachment-specific) if there are new installations or proposed changes
        midspan_proposed_value = extract_string_value(pole.get('midspan_proposed'))
        
        # Only show midspan_proposed if it's not 'N/A'
        if midspan_proposed_value != 'N/A':
            proposed_midspan_cell.value = midspan_proposed_value
        else:
            proposed_midspan_cell.value = ''
            
        proposed_midspan_cell.alignment = center_alignment
        proposed_midspan_cell.border = thin_border
        
        # Process REF Groups for this primary pole
        current_row = midspan_row + 1
        
        # ===== Write REF Groups =====
        for ref_group in ref_groups:
            # REF Group Header
            ref_header_row = current_row
            
            # Determine fill color based on style_hint and label
            ref_fill = orange_fill  # Default to orange
            ref_label = extract_string_value(ref_group.get('label', ''))
            
            if ref_group.get('style_hint') == 'purple':
                ref_fill = purple_fill
            elif ref_group.get('style_hint') == 'light-blue':
                ref_fill = light_blue_fill
            # Also check the label for South East references
            elif 'south east' in ref_label.lower() or 'southeast' in ref_label.lower():
                ref_fill = purple_fill
            
            # We need to place the header in column L (12), not column J (10)
            # This matches the expected image where reference spans appear in the Make Ready Data section
            ref_header_cell = ws.cell(row=ref_header_row, column=12)
            ref_header_cell.value = ref_label
            
            # Apply special formatting
            ref_header_cell.fill = ref_fill
            ref_header_cell.font = header_font
            ref_header_cell.alignment = center_alignment
            ref_header_cell.border = thin_border
            
            # Merge cells L-O for header (columns 12-15)
            ws.merge_cells(start_row=ref_header_row, start_column=12, 
                          end_row=ref_header_row, end_column=15)
            
            # Add borders to all merged cells
            for col in range(12, 16):
                cell = ws.cell(row=ref_header_row, column=col)
                cell.border = thin_border
            
            # Move to next row for wire data
            current_row += 1
            
            # Try to use attachments from the reference span data first
            attachments = ref_group.get('attachments', [])
            
            if attachments:
                # Use pre-processed attachments if available
                for att in attachments:
                    # Attacher description
                    desc_cell = ws.cell(row=current_row, column=12)
                    desc_cell.value = extract_string_value(att.get('description'))
                    desc_cell.alignment = Alignment(horizontal='left', vertical='center')
                    desc_cell.border = thin_border
                    
                    # Existing height
                    existing_cell = ws.cell(row=current_row, column=13)
                    existing_cell.value = extract_string_value(att.get('existing_height'))
                    existing_cell.alignment = center_alignment
                    existing_cell.border = thin_border
                    
                    # Proposed height
                    proposed_cell = ws.cell(row=current_row, column=14)
                    proposed_cell.value = extract_string_value(att.get('proposed_height'))
                    proposed_cell.alignment = center_alignment
                    proposed_cell.border = thin_border
                    
                    # Midspan proposed
                    midspan_cell = ws.cell(row=current_row, column=15)
                    midspan_cell.value = extract_string_value(att.get('midspan_proposed'))
                    midspan_cell.alignment = center_alignment
                    midspan_cell.border = thin_border
                    
                    current_row += 1
            else:
                # Fall back to connection_heights data
                connection_id = ref_group.get('connection_id')
                if connection_id in connection_heights:
                    for wire_type, height_data in connection_heights[connection_id].items():
                        # Wire description in column L
                        desc_cell = ws.cell(row=current_row, column=12)
                        desc_cell.value = extract_string_value(wire_type)
                        desc_cell.alignment = Alignment(horizontal='left', vertical='center')
                        desc_cell.border = thin_border
                        
                        # Existing height in column M
                        existing_cell = ws.cell(row=current_row, column=13)
                        existing_cell.value = extract_string_value(height_data.get('height_str'))
                        existing_cell.alignment = center_alignment
                        existing_cell.border = thin_border
                        
                        # Proposed height in column N
                        proposed_cell = ws.cell(row=current_row, column=14)
                        proposed_cell.alignment = center_alignment
                        proposed_cell.border = thin_border
                        
                        # Check if there's a proposed height
                        has_proposed = (connection_id in proposed_heights and 
                                      wire_type in proposed_heights[connection_id])
                        
                        if has_proposed:
                            # Show proposed height
                            proposed_cell.value = extract_string_value(proposed_heights[connection_id][wire_type].get('height_str'))
                        
                        # Midspan proposed in column O
                        midspan_cell = ws.cell(row=current_row, column=15)
                        midspan_cell.alignment = center_alignment
                        midspan_cell.border = thin_border
                        
                        # Try to determine an appropriate midspan value based on wire type
                        midspan_height_in = None
                        if 'midspanHeight_in' in height_data['wire_data']:
                            midspan_height_in = height_data['wire_data'].get('midspanHeight_in')
                            midspan_cell.value = inches_to_feet_inches_str(midspan_height_in)
                        
                        current_row += 1
        
        # Add blank row to separate poles
        for col in range(1, 16):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
        
        current_row += 1
        operation_number += 1  # Increment only for primary poles

    # Save the workbook
    wb.save(output_filepath)
    print(f"[INFO] Make Ready Report saved to {output_filepath}")

# If you need to test the function independently
if __name__ == "__main__":
    # Example data for testing
    test_data = [
        {
            "pole_owner": "CPS ENERGY",
            "pole_number": "PL410620",
            "pole_structure": "40-4 Southern Pine",
            "proposed_riser": "No",
            "proposed_guy": "No",
            "pla_percentage": "78.70%",
            "construction_grade": "C",
            "existing_midspan_lowest_com": "23'10\"",
            "existing_midspan_lowest_cps_electrical": "29'6\"",
            "midspan_proposed": "21'4\"",
            "from_pole": "PL410620",
            "to_pole": "PL398491",
            "connections": [
                {
                    "connection_id": "conn-1",
                    "from_pole": "PL410620",
                    "to_pole": "PL398491",
                    "lowest_com": 286,
                    "lowest_cps": 354
                }
            ],
            "attachers": [
                {
                    "description": "Neutral",
                    "existing_height": "29'4\"",
                    "proposed_height": "28'6\"",
                    "midspan_proposed": "N/A"
                },
                {
                    "description": "CPS Supply Fiber",
                    "existing_height": "28'6\"",
                    "proposed_height": "N/A",
                    "midspan_proposed": "N/A"
                },
                {
                    "description": "Charter Spectrum Fiber Optic",
                    "existing_height": "27'7\"",
                    "proposed_height": "26'7\"",
                    "midspan_proposed": "21'4\""
                },
                {
                    "description": "AT&T Fiber Optic Com",
                    "existing_height": "25'7\"",
                    "proposed_height": "N/A",
                    "midspan_proposed": "N/A"
                },
                {
                    "description": "AT&T Telco Com",
                    "existing_height": "22'4\"",
                    "proposed_height": "N/A",
                    "midspan_proposed": "N/A"
                },
                {
                    "description": "AT&T Com Drop",
                    "existing_height": "21'5\"",
                    "proposed_height": "15'10\"",
                    "midspan_proposed": "N/A"
                }
            ]
        }
    ]
    
    create_make_ready_excel(test_data, "test_output.xlsx")
